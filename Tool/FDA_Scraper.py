#!/usr/bin/env python3
import pandas as pd
import settings
import time
import os
import requests
import urllib.parse
import re
from openpyxl import load_workbook
from settings import logger
from urllib.parse import urljoin, urlparse

def load_target_submission_ids(input_file_path: str) -> set:
    """Load target submission IDs from input.txt file"""
    target_ids = set()
    try:
        with open(input_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                submission_id = line.strip()
                if submission_id:
                    target_ids.add(submission_id)
        logger.info(f"Loaded {len(target_ids)} target submission IDs from {input_file_path}")
        return target_ids
    except Exception as e:
        logger.error(f"Error loading target IDs from {input_file_path}: {str(e)}")
        return set()

def extract_hyperlinks_from_excel(file_path: str, sheet_name: str = None) -> dict:
    """Extract submission numbers and construct FDA URLs directly"""
    logger.info("Constructing FDA URLs from submission numbers in Column B")
    
    try:
        wb = load_workbook(file_path)
        ws = wb[sheet_name] if sheet_name else wb.active
        hyperlinks = {}
        
        for row in range(2, min(ws.max_row + 1, 1000)):
            cell = ws[f'B{row}']
            
            if cell.value:
                submission_number = str(cell.value).strip()
                
                if not submission_number or submission_number.lower() in ['nan', 'none', '']:
                    continue
                
                fda_url = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfPMN/pmn.cfm?ID={submission_number}"
                hyperlinks[submission_number] = fda_url
                logger.info(f"CONSTRUCTED: {submission_number} -> {fda_url}")
        
        logger.info(f"Constructed {len(hyperlinks)} URLs from submission numbers")
        return hyperlinks
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {str(e)}")
        return {}

def extract_url_from_hyperlink_formula(formula_text: str) -> tuple:
    """Extract URL and display text from Excel HYPERLINK formula"""
    try:
        patterns = [
            r'=HYPERLINK\("([^"]+)",\s*"([^"]+)"\)',
            r'=HYPERLINK\(\"([^\"]+)\",\s*\"([^\"]+)\"\)',
            r'=HYPERLINK\(([^,]+),\s*([^)]+)\)'
        ]
        
        for pattern in patterns:
            match = re.match(pattern, formula_text)
            if match:
                url = match.group(1).strip('"\'')
                display_text = match.group(2).strip('"\'')
                return url, display_text
                
        return None, None
    except Exception as e:
        logger.error(f"Error extracting from HYPERLINK formula: {str(e)}")
        return None, None

def construct_fda_urls_from_submission_numbers(file_path: str) -> dict:
    """Construct FDA URLs directly from submission numbers"""
    try:
        logger.info("Constructing FDA URLs from submission numbers")
        wb = load_workbook(file_path)
        ws = wb.active
        hyperlinks = {}
        
        for row in range(2, min(ws.max_row + 1, 1000)):
            cell = ws[f'B{row}']
            if cell.value:
                submission_number = str(cell.value).strip()
                
                if not submission_number or submission_number.lower() in ['nan', 'none', '']:
                    continue
                
                fda_url = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfPMN/pmn.cfm?ID={submission_number}"
                hyperlinks[submission_number] = fda_url
                logger.info(f"CONSTRUCTED: {submission_number} -> {fda_url}")
        
        logger.info(f"Constructed {len(hyperlinks)} URLs from submission numbers")
        return hyperlinks
        
    except Exception as e:
        logger.error(f"Error constructing URLs: {str(e)}")
        return {}

def generate_fda_url_for_submission(submission_number: str) -> str:
    """Generate FDA URL using specified pattern"""
    return f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfPMN/pmn.cfm?ID={submission_number}"

def find_pdf_links_in_html(html_content: str, base_url: str) -> list:
    """Find PDF links in HTML content"""
    pdf_links = []
    
    patterns = [
        r'href=["\']([^"\']*\.pdf[^"\']*)["\']',
        r'href=["\']([^"\']*[Ss]ummary[^"\']*)["\']',
        r'href=["\']([^"\']*cfmis[^"\']*[Ss]ummary[^"\']*)["\']',
        r'href=["\']([^"\']*[Dd]ocument[^"\']*)["\']',
        r'href=["\']([^"\']*cfmis[^"\']*)["\']',
        r'href=["\']([^"\']*cfdocs[^"\']*pdf[^"\']*)["\']',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, html_content, re.IGNORECASE)
        for match in matches:
            if not match.startswith('http'):
                match = urljoin(base_url, match)
            pdf_links.append(match)
    
    summary_text_patterns = [
        r'<a[^>]*href=["\']([^"\']+)["\'][^>]*>Summary</a>',
        r'<a[^>]*href=["\']([^"\']+)["\'][^>]*>[^<]*Summary[^<]*</a>',
    ]
    
    for pattern in summary_text_patterns:
        matches = re.findall(pattern, html_content, re.IGNORECASE)
        for match in matches:
            if not match.startswith('http'):
                match = urljoin(base_url, match)
            pdf_links.append(match)
            logger.info(f"Found FDA Summary link: {match}")
    
    seen = set()
    unique_links = []
    for link in pdf_links:
        if link not in seen:
            seen.add(link)
            unique_links.append(link)
    
    return unique_links

def find_fda_summary_links(html_content: str, base_url: str) -> list:
    """Find FDA Summary links"""
    summary_links = []
    
    patterns = [
        r'<a[^>]*href=["\']([^"\']*)["\'][^>]*>Summary</a>',
        r'<td[^>]*><a[^>]*href=["\']([^"\']*)["\'][^>]*>Summary</a></td>',
        r'href=["\']([^"\']*cfmis[^"\']*view=1[^"\']*)["\']',
        r'href=["\']([^"\']*cfmis[^"\']*Summary[^"\']*)["\']',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, html_content, re.IGNORECASE)
        for match in matches:
            if not match.startswith('http'):
                match = urljoin(base_url, match)
            summary_links.append(match)
            logger.info(f"Found FDA Summary link: {match}")
    
    return summary_links

def download_pdf_requests_only(page_url: str, filename: str, max_retries: int = 3) -> bool:
    """Download PDF using requests only"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Referer': 'https://www.accessdata.fda.gov/',
    }
    
    session = requests.Session()
    session.headers.update(headers)
    
    for attempt in range(max_retries):
        try:
            logger.info(f"PROCESSING: {page_url} (Attempt {attempt + 1})")
            
            if attempt > 0:
                time.sleep(5 * attempt)
            
            response = session.get(page_url, timeout=30, allow_redirects=True)
            response.raise_for_status()
            
            content_type = response.headers.get('content-type', '').lower()
            if 'application/pdf' in content_type or response.content.startswith(b'%PDF'):
                logger.info("Page directly returns PDF content")
                return save_pdf_content(response.content, filename)
            
            html_content = response.text
            pdf_links = find_pdf_links_in_html(html_content, page_url)
            
            logger.info(f"Found {len(pdf_links)} potential PDF links")
            
            summary_links = find_fda_summary_links(html_content, page_url)
            if summary_links:
                logger.info(f"Found {len(summary_links)} FDA Summary links")
                pdf_links = summary_links + pdf_links
            
            if not pdf_links:
                submission_id = extract_submission_id_from_url(page_url)
                if submission_id:
                    pdf_links = generate_common_fda_pdf_urls(page_url, submission_id)
                    logger.info(f"Generated {len(pdf_links)} common FDA PDF URLs")
            
            for i, pdf_url in enumerate(pdf_links):
                try:
                    logger.info(f"Trying PDF link {i+1}/{len(pdf_links)}: {pdf_url}")
                    
                    pdf_response = session.get(pdf_url, timeout=45, allow_redirects=True)
                    pdf_response.raise_for_status()
                    
                    if 'text/html' in pdf_response.headers.get('content-type', '').lower():
                        logger.info("Got HTML response, looking for PDF links...")
                        summary_pdf_links = find_pdf_links_in_html(pdf_response.text, pdf_url)
                        
                        for summary_pdf_url in summary_pdf_links:
                            try:
                                logger.info(f"Trying summary PDF: {summary_pdf_url}")
                                summary_pdf_response = session.get(summary_pdf_url, timeout=45, allow_redirects=True)
                                summary_pdf_response.raise_for_status()
                                
                                if summary_pdf_response.content.startswith(b'%PDF'):
                                    logger.info(f"SUCCESS: Found valid PDF at {summary_pdf_url}")
                                    return save_pdf_content(summary_pdf_response.content, filename)
                            except:
                                continue
                    
                    elif pdf_response.content.startswith(b'%PDF'):
                        logger.info(f"SUCCESS: Found valid PDF at {pdf_url}")
                        return save_pdf_content(pdf_response.content, filename)
                        
                except requests.exceptions.RequestException as e:
                    logger.debug(f"Failed to download {pdf_url}: {str(e)}")
                    continue
                except Exception as e:
                    logger.debug(f"Unexpected error with {pdf_url}: {str(e)}")
                    continue
            
            if attempt == max_retries - 1:
                logger.warning("No PDFs found, saving HTML as fallback")
                return save_html_as_fallback(html_content, filename.replace('.pdf', '.html'))
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Request error for {page_url} (Attempt {attempt + 1}): {str(e)}")
        except Exception as e:
            logger.error(f"Unexpected error for {page_url} (Attempt {attempt + 1}): {str(e)}")
    
    return False

def save_pdf_content(pdf_content: bytes, filename: str) -> bool:
    """Save PDF content to file"""
    try:
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        with open(filename, 'wb') as f:
            f.write(pdf_content)
        
        file_size = os.path.getsize(filename)
        if file_size > 1000:
            logger.info(f"Successfully saved PDF: {filename} ({file_size:,} bytes)")
            return True
        else:
            logger.error(f"Generated PDF is too small: {filename} ({file_size} bytes)")
            if os.path.exists(filename):
                os.remove(filename)
            return False
            
    except Exception as e:
        logger.error(f"Error saving PDF: {str(e)}")
        return False

def save_html_as_fallback(html_content: str, filename: str) -> bool:
    """Save HTML content as fallback"""
    try:
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        file_size = os.path.getsize(filename)
        logger.info(f"Saved HTML fallback: {filename} ({file_size:,} bytes)")
        return True
        
    except Exception as e:
        logger.error(f"Error saving HTML fallback: {str(e)}")
        return False

def extract_submission_id_from_url(url: str) -> str:
    """Extract submission ID from FDA URL"""
    patterns = [
        r'[ID|id]=([PK]\d+)',
        r'/([PK]\d+)$',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    
    return None

def generate_common_fda_pdf_urls(base_url: str, submission_id: str) -> list:
    """Generate common FDA PDF URL patterns"""
    base_domain = "https://www.accessdata.fda.gov"
    pdf_urls = []
    
    year_match = re.search(r'(\d{2})(\d+)$', submission_id)
    if year_match:
        year_part = year_match.group(1)
        
        patterns = [
            f"{base_domain}/cdrh_docs/pdf{year_part}/{submission_id}.pdf",
            f"{base_domain}/cdrh_docs/pdf{year_part}/{submission_id}_summary.pdf",
            f"{base_domain}/cdrh_docs/pdf{year_part}/{submission_id}a000.pdf",
            f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmis/pmn.cfm?ID={submission_id}&view=1",
        ]
        
        pdf_urls.extend(patterns)
    
    pdf_urls.append(f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmis/pmn.cfm?ID={submission_id}")
    
    return pdf_urls

def process_single_submission(submission_data: tuple) -> tuple:
    """Process a single submission"""
    submission_number, page_url, pdf_filename = submission_data
    
    logger.info(f"STARTING: {submission_number}")
    
    if os.path.exists(pdf_filename) and os.path.getsize(pdf_filename) > 1000:
        logger.info(f"SKIPPED: {submission_number} - File already exists")
        return ('skipped', submission_number)
    
    success = download_pdf_requests_only(page_url, pdf_filename)
    
    if success:
        logger.info(f"COMPLETED: {submission_number}")
        return ('success', submission_number)
    else:
        logger.error(f"FAILED: {submission_number}")
        return ('failed', submission_number)

def download_reports() -> None:
    """Main function for downloading PDFs"""
    logger.info("Starting PDF downloads from Excel submission numbers")
    
    input_file_path = os.path.join("mnt", "Data", "input.txt")
    target_ids = load_target_submission_ids(input_file_path)
    
    if not target_ids:
        logger.error("No target submission IDs found. Exiting.")
        return
    
    try:
        data = pd.read_excel(settings.EXCEL_FILE)
        logger.info(f"Loaded Excel file with {len(data)} rows")
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        return
    
    pdf_dir = os.path.join("mnt", "Data", "Summary_docs")
    os.makedirs(pdf_dir, exist_ok=True)
    logger.info(f"Download directory: {os.path.abspath(pdf_dir)}")
    
    successful_downloads = 0
    failed_downloads = 0
    skipped_existing = 0
    filtered_out = 0
    
    process_limit = getattr(settings, 'PROCESS_LIMIT', 10)
    processed = 0
    
    logger.info(f"Processing up to {process_limit} submissions that match target IDs")
    
    for index, row in data.iterrows():
        if processed >= process_limit:
            break
            
        try:
            submission_number = str(row["Submission Number"]).strip()
            
            if not submission_number or submission_number.lower() in ['nan', 'none', '']:
                logger.debug(f"Skipping invalid submission number: {submission_number}")
                continue
            
            if submission_number not in target_ids:
                filtered_out += 1
                logger.debug(f"FILTERED OUT: {submission_number} - Not in target list")
                continue
            
            pdf_filename = os.path.join(pdf_dir, f"{submission_number}.pdf")
            
            page_url = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfPMN/pmn.cfm?ID={submission_number}"
            logger.info(f"PROCESSING {processed+1}/{process_limit}: {submission_number}")
            logger.info(f"URL: {page_url}")
            
            result_type, result_submission = process_single_submission((submission_number, page_url, pdf_filename))
            
            if result_type == 'success':
                successful_downloads += 1
            elif result_type == 'skipped':
                skipped_existing += 1
            else:
                failed_downloads += 1
            
            logger.info(f"PROGRESS: {processed+1}/{process_limit} | SUCCESS:{successful_downloads} FAILURE:{failed_downloads} SKIPPED:{skipped_existing}")
            
            processed += 1
            
            if processed < process_limit:
                logger.info("Waiting 3 seconds...")
                time.sleep(3)
            
        except Exception as e:
            logger.error(f"Error processing submission {submission_number}: {str(e)}")
            failed_downloads += 1
            processed += 1
            continue
    
    logger.info("="*60)
    logger.info("FINAL DOWNLOAD SUMMARY:")
    logger.info(f"Total submissions in Excel: {len(data)}")
    logger.info(f"Filtered out (not in target list): {filtered_out}")
    logger.info(f"Total processed: {processed}")
    logger.info(f"Successful downloads: {successful_downloads}")
    logger.info(f"Failed downloads: {failed_downloads}")
    logger.info(f"Skipped (already exist): {skipped_existing}")
    logger.info(f"Success rate: {(successful_downloads/(processed-skipped_existing)*100):.1f}%" if processed > skipped_existing else "N/A")
    logger.info("="*60)

if __name__ == "__main__":
    try:
        logger.info("Starting download process...")
        
        if not os.path.exists(settings.EXCEL_FILE):
            logger.error(f"Excel file not found: {settings.EXCEL_FILE}")
            logger.error("Please ensure the Excel file exists before running the script.")
        else:
            logger.info(f"Excel file found: {settings.EXCEL_FILE}")
            download_reports()

    except KeyboardInterrupt:
        logger.critical("Raised KeyboardInterrupt -> Closing script...")
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")